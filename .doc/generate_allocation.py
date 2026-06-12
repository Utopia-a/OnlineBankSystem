# -*- coding: utf-8 -*-
"""生成 10 人代码分配 Excel 文档"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MEMBERS = [
    {
        "id": "A", "name": "成员A", "type": "原核心成员", "role": "前端负责人",
        "duty": "认证相关页面、全局样式与 JS、页面路由、Swagger 文档配置",
        "files": [
            ("前端页面", "project/front/templates/login.html", "登录页"),
            ("前端页面", "project/front/templates/register.html", "注册页（含邮箱验证步骤）"),
            ("前端页面", "project/front/templates/verify-email.html", "邮箱验证页"),
            ("前端页面", "project/front/templates/dashboard.html", "首页 / 账户概览"),
            ("前端页面", "project/front/templates/fragments/nav.html", "公共导航栏片段"),
            ("静态资源", "project/front/static/css/common.css", "全局样式（设计系统）"),
            ("静态资源", "project/front/static/js/api.js", "HTTP 请求封装（JWT、错误处理）"),
            ("静态资源", "project/front/static/js/app.js", "公共工具（金额、日期、Toast）"),
            ("静态资源", "project/front/static/js/auth-verify.js", "邮箱验证相关 API"),
            ("文档", "project/front/README.md", "前端目录说明"),
            ("后端-路由", "project/back/src/main/java/com/bank/controller/PageController.java", "Thymeleaf 页面路由"),
            ("后端-配置", "project/back/src/main/java/com/bank/config/WebMvcConfig.java", "静态资源、CORS、首页重定向"),
            ("后端-配置", "project/back/src/main/java/com/banking/config/OpenApiConfig.java", "Swagger / OpenAPI 配置"),
            ("配置", "project/back/src/main/resources/application.yml", "springdoc.* Swagger 路径（协同 I 维护）"),
        ],
    },
    {
        "id": "B", "name": "成员B", "type": "原核心成员", "role": "认证安全工程师",
        "duty": "注册/登录、JWT、Spring Security、OTP 邮箱验证、用户表",
        "files": [
            ("DTO", "project/back/src/main/java/com/banking/auth/dto/AuthRequest.java", "登录、注册、OTP、改密请求"),
            ("DTO", "project/back/src/main/java/com/banking/auth/dto/AuthResponse.java", "登录结果、Token 响应"),
            ("实体", "project/back/src/main/java/com/banking/auth/entity/User.java", "用户实体"),
            ("实体", "project/back/src/main/java/com/banking/auth/entity/RefreshToken.java", "刷新 Token 实体"),
            ("实体", "project/back/src/main/java/com/banking/auth/entity/OtpRecord.java", "OTP 记录实体"),
            ("异常", "project/back/src/main/java/com/banking/auth/exception/AuthException.java", "认证异常"),
            ("过滤器", "project/back/src/main/java/com/banking/auth/filter/JwtAuthenticationFilter.java", "JWT 过滤器"),
            ("Repository", "project/back/src/main/java/com/banking/auth/repository/UserRepository.java", "用户仓储"),
            ("Repository", "project/back/src/main/java/com/banking/auth/repository/RefreshTokenRepository.java", "Token 仓储"),
            ("Repository", "project/back/src/main/java/com/banking/auth/repository/OtpRecordRepository.java", "OTP 仓储"),
            ("Service", "project/back/src/main/java/com/banking/auth/service/AuthService.java", "认证核心逻辑"),
            ("Service", "project/back/src/main/java/com/banking/auth/service/OtpService.java", "OTP 生成与校验"),
            ("Service", "project/back/src/main/java/com/banking/auth/service/RefreshTokenService.java", "Refresh Token 管理"),
            ("Service", "project/back/src/main/java/com/banking/auth/service/CustomUserDetailsService.java", "用户加载"),
            ("工具", "project/back/src/main/java/com/banking/auth/util/JwtUtil.java", "JWT 签发与解析"),
            ("工具", "project/back/src/main/java/com/banking/auth/util/OtpUtil.java", "OTP 工具"),
            ("配置", "project/back/src/main/java/com/banking/config/SecurityConfig.java", "Security 白名单与过滤器链"),
            ("共享", "project/back/src/main/java/com/bank/account/config/BankUserDetails.java", "UserDetails 扩展（协同 C）"),
            ("配置", "project/back/src/main/resources/application.yml", "jwt.*、otp.* 配置节（协同 I）"),
            ("数据库", "project/back/sql/schema.sql", "users / refresh_tokens / otp_records 表段（协同 J）"),
            ("测试", "project/back/src/test/java/com/banking/auth/AuthServiceTest.java", "认证服务测试"),
            ("测试", "project/back/src/test/java/com/banking/auth/JwtUtilTest.java", "JWT 工具测试"),
            ("测试", "project/back/src/test/java/com/banking/auth/AdminPasswordHashTest.java", "管理员密码哈希测试"),
        ],
    },
    {
        "id": "C", "name": "成员C", "type": "原核心成员", "role": "账户模块开发",
        "duty": "账户 CRUD、余额查询、账户状态、内部服务接口",
        "files": [
            ("配置", "project/back/src/main/java/com/bank/account/config/AccountNumberGenerator.java", "账户号生成器"),
            ("Controller", "project/back/src/main/java/com/bank/account/controller/AdminAccountController.java", "管理员账户接口"),
            ("Controller", "project/back/src/main/java/com/bank/account/controller/InternalAccountController.java", "内部服务接口"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/ApiResponse.java", "模块响应包装"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/AccountResponse.java", "账户响应"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/BalanceResponse.java", "余额响应"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/CreateAccountRequest.java", "开户请求"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/UpdateAccountRequest.java", "更新账户请求"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/AccountStatusRequest.java", "状态变更请求"),
            ("DTO", "project/back/src/main/java/com/bank/account/dto/InternalBalanceUpdateRequest.java", "内部余额更新"),
            ("实体", "project/back/src/main/java/com/bank/account/entity/Account.java", "账户实体"),
            ("枚举", "project/back/src/main/java/com/bank/account/enums/AccountType.java", "账户类型"),
            ("枚举", "project/back/src/main/java/com/bank/account/enums/AccountStatus.java", "账户状态"),
            ("异常", "project/back/src/main/java/com/bank/account/exception/AccountNotFoundException.java", "账户不存在"),
            ("异常", "project/back/src/main/java/com/bank/account/exception/AccountStatusException.java", "账户状态异常"),
            ("异常", "project/back/src/main/java/com/bank/account/exception/AccountAccessDeniedException.java", "访问拒绝"),
            ("异常", "project/back/src/main/java/com/bank/account/exception/InsufficientBalanceException.java", "余额不足"),
            ("Repository", "project/back/src/main/java/com/bank/account/repository/AccountRepository.java", "账户仓储"),
            ("Service", "project/back/src/main/java/com/bank/account/service/AccountService.java", "账户服务接口"),
            ("Service", "project/back/src/main/java/com/bank/account/service/AccountServiceImpl.java", "账户服务实现"),
            ("数据库", "project/back/sql/schema.sql", "accounts 表段（协同 J）"),
            ("测试", "project/back/src/test/java/com/bank/account/AccountServiceTest.java", "账户服务测试"),
        ],
    },
    {
        "id": "D", "name": "成员D", "type": "原核心成员", "role": "交易模块开发",
        "duty": "转账、存款、取款、事务控制、流水号生成",
        "files": [
            ("配置", "project/back/src/main/java/com/bank/transaction/config/TransactionNoGenerator.java", "流水号生成"),
            ("配置", "project/back/src/main/java/com/bank/transaction/config/TransactionProperties.java", "限额配置"),
            ("DTO", "project/back/src/main/java/com/bank/transaction/dto/ApiResponse.java", "模块响应包装"),
            ("DTO", "project/back/src/main/java/com/bank/transaction/dto/DepositRequest.java", "存款请求"),
            ("DTO", "project/back/src/main/java/com/bank/transaction/dto/WithdrawRequest.java", "取款请求"),
            ("DTO", "project/back/src/main/java/com/bank/transaction/dto/TransferRequest.java", "转账请求"),
            ("DTO", "project/back/src/main/java/com/bank/transaction/dto/TransactionResponse.java", "交易响应"),
            ("实体", "project/back/src/main/java/com/bank/transaction/entity/Transaction.java", "交易实体"),
            ("枚举", "project/back/src/main/java/com/bank/transaction/enums/TransactionType.java", "交易类型"),
            ("枚举", "project/back/src/main/java/com/bank/transaction/enums/TransactionStatus.java", "交易状态"),
            ("异常", "project/back/src/main/java/com/bank/transaction/exception/BusinessException.java", "通用业务异常"),
            ("异常", "project/back/src/main/java/com/bank/transaction/exception/InsufficientBalanceException.java", "余额不足"),
            ("异常", "project/back/src/main/java/com/bank/transaction/exception/AccountNotFoundException.java", "账户不存在"),
            ("异常", "project/back/src/main/java/com/bank/transaction/exception/AccountStatusException.java", "账户状态异常"),
            ("异常", "project/back/src/main/java/com/bank/transaction/exception/TransactionLimitException.java", "交易限额"),
            ("Repository", "project/back/src/main/java/com/bank/transaction/repository/TransactionRepository.java", "交易仓储"),
            ("Service", "project/back/src/main/java/com/bank/transaction/service/TransactionService.java", "交易服务接口"),
            ("Service", "project/back/src/main/java/com/bank/transaction/service/TransactionServiceImpl.java", "交易服务实现"),
            ("配置", "project/back/src/main/resources/application.yml", "transaction.* 限额配置节（协同 I）"),
            ("数据库", "project/back/sql/schema.sql", "transactions 表段（协同 J）"),
            ("测试", "project/back/src/test/java/com/bank/transaction/TransactionServiceTest.java", "交易服务测试"),
        ],
    },
    {
        "id": "E", "name": "成员E", "type": "原核心成员", "role": "报表模块开发",
        "duty": "交易历史查询、分页、统计、CSV/Excel/PDF 导出",
        "files": [
            ("DTO", "project/back/src/main/java/com/banking/report/dto/ReportResponse.java", "报表响应 DTO 集合"),
            ("DTO", "project/back/src/main/java/com/banking/report/dto/TransactionQueryRequest.java", "查询条件"),
            ("异常", "project/back/src/main/java/com/banking/report/exception/ReportException.java", "报表异常"),
            ("Repository", "project/back/src/main/java/com/banking/report/repository/TransactionSpecification.java", "动态查询"),
            ("Service", "project/back/src/main/java/com/banking/report/service/ReportService.java", "报表核心服务"),
            ("Service", "project/back/src/main/java/com/banking/report/service/UserContext.java", "用户上下文"),
            ("工具", "project/back/src/main/java/com/banking/report/util/ExportUtil.java", "CSV/Excel/PDF 导出"),
            ("工具", "project/back/src/main/java/com/banking/report/util/PageUtil.java", "分页与排序安全"),
            ("配置", "project/back/src/main/resources/application.yml", "export.* 配置节（协同 I）"),
            ("测试", "project/back/src/test/java/com/banking/report/ExportUtilTest.java", "导出工具测试"),
        ],
    },
    {
        "id": "F", "name": "成员F", "type": "原核心成员", "role": "管理员后台开发",
        "duty": "用户管理、交易监控、系统配置、管理员数据初始化",
        "files": [
            ("Controller", "project/back/src/main/java/com/bank/admin/controller/AdminUserController.java", "用户管理 API"),
            ("Controller", "project/back/src/main/java/com/bank/admin/controller/TransactionMonitorController.java", "交易监控 API"),
            ("Controller", "project/back/src/main/java/com/bank/admin/controller/SystemConfigController.java", "系统配置 API"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/request/PageRequest.java", "分页请求"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/request/UserQueryRequest.java", "用户查询"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/request/UpdateUserStatusRequest.java", "状态更新"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/request/TransactionQueryRequest.java", "交易查询"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/request/SystemConfigRequest.java", "配置请求"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/response/Result.java", "统一响应"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/response/PageResult.java", "分页响应"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/response/UserVO.java", "用户 VO"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/response/TransactionVO.java", "交易 VO"),
            ("DTO", "project/back/src/main/java/com/bank/admin/dto/response/DashboardStatsVO.java", "仪表盘统计"),
            ("实体", "project/back/src/main/java/com/bank/admin/entity/SystemConfig.java", "系统配置实体"),
            ("枚举", "project/back/src/main/java/com/bank/admin/enums/AdminTransactionType.java", "管理端交易类型"),
            ("Repository", "project/back/src/main/java/com/bank/admin/repository/AdminUserRepository.java", "用户仓储"),
            ("Repository", "project/back/src/main/java/com/bank/admin/repository/AdminTransactionRepository.java", "交易仓储"),
            ("Repository", "project/back/src/main/java/com/bank/admin/repository/SystemConfigRepository.java", "配置仓储"),
            ("Service", "project/back/src/main/java/com/bank/admin/service/AdminUserService.java", "用户管理服务"),
            ("Service", "project/back/src/main/java/com/bank/admin/service/TransactionMonitorService.java", "交易监控服务"),
            ("Service", "project/back/src/main/java/com/bank/admin/service/SystemConfigService.java", "配置服务"),
            ("Service", "project/back/src/main/java/com/bank/admin/service/impl/AdminUserServiceImpl.java", "用户管理实现"),
            ("Service", "project/back/src/main/java/com/bank/admin/service/impl/TransactionMonitorServiceImpl.java", "交易监控实现"),
            ("Service", "project/back/src/main/java/com/bank/admin/service/impl/SystemConfigServiceImpl.java", "配置服务实现"),
            ("数据库", "project/back/sql/schema.sql", "system_config 表及 admin 账号（协同 J）"),
            ("数据库", "project/back/sql/fix-admin-password.sql", "修复管理员密码"),
            ("测试", "project/back/src/test/java/com/bank/admin/service/AdminUserServiceTest.java", "用户管理测试"),
            ("测试", "project/back/src/test/java/com/bank/admin/service/SystemConfigServiceTest.java", "配置服务测试"),
        ],
    },
    {
        "id": "G", "name": "成员G", "type": "新增成员", "role": "API 集成工程师",
        "duty": "Web API 控制器、前端 DTO 适配、全局异常处理",
        "files": [
            ("Controller", "project/back/src/main/java/com/bank/web/WebAuthController.java", "/api/auth/** 认证接口"),
            ("Controller", "project/back/src/main/java/com/bank/web/WebAccountController.java", "/api/accounts/** 账户接口"),
            ("Controller", "project/back/src/main/java/com/bank/web/WebTransactionController.java", "/api/transactions/** 交易接口"),
            ("Controller", "project/back/src/main/java/com/bank/web/WebBillController.java", "/api/bills/** 账单接口"),
            ("DTO", "project/back/src/main/java/com/bank/dto/request/LoginRequest.java", "登录请求"),
            ("DTO", "project/back/src/main/java/com/bank/dto/request/RegisterRequest.java", "注册请求"),
            ("DTO", "project/back/src/main/java/com/bank/dto/request/TransferRequest.java", "转账请求"),
            ("DTO", "project/back/src/main/java/com/bank/dto/request/DepositWithdrawRequest.java", "存取款请求"),
            ("DTO", "project/back/src/main/java/com/bank/dto/response/ApiResponse.java", "统一 API 响应"),
            ("DTO", "project/back/src/main/java/com/bank/dto/response/LoginResponse.java", "登录响应"),
            ("DTO", "project/back/src/main/java/com/bank/dto/response/RegisterResponse.java", "注册响应"),
            ("DTO", "project/back/src/main/java/com/bank/dto/response/AccountResponse.java", "账户响应"),
            ("DTO", "project/back/src/main/java/com/bank/dto/response/TransactionResponse.java", "交易响应"),
            ("DTO", "project/back/src/main/java/com/bank/dto/response/PageResponse.java", "分页响应"),
            ("配置", "project/back/src/main/java/com/banking/config/GlobalExceptionHandler.java", "全局异常处理"),
        ],
    },
    {
        "id": "H", "name": "成员H", "type": "新增成员", "role": "业务前端开发",
        "duty": "账户、转账、交易记录、个人中心等业务页面",
        "files": [
            ("前端页面", "project/front/templates/accounts.html", "账户管理页"),
            ("前端页面", "project/front/templates/transfer.html", "转账汇款页"),
            ("前端页面", "project/front/templates/transactions.html", "交易记录页（含导出）"),
            ("前端页面", "project/front/templates/profile.html", "个人中心页"),
        ],
    },
    {
        "id": "I", "name": "成员I", "type": "新增成员", "role": "基础设施工程师",
        "duty": "工程构建、启动类、运行环境、全局配置统筹",
        "files": [
            ("启动", "project/back/src/main/java/com/banking/OnlineBankApplication.java", "Spring Boot 启动类"),
            ("构建", "project/back/pom.xml", "后端 Maven 配置（含 front 资源打包）"),
            ("构建", "pom.xml", "根目录 Maven 聚合工程"),
            ("配置", "project/back/src/main/resources/application.yml", "数据库、服务端口等全局配置（主维护）"),
            ("运维", "project/back/docker-compose.yml", "MySQL 本地环境"),
            ("脚本", "project/back/scripts/start.bat", "一键启动"),
            ("脚本", "project/back/scripts/build-all.bat", "编译脚本"),
            ("文档", "README.md", "项目说明"),
            ("文档", ".doc/SETUP.md", "部署与 IDEA 配置"),
            ("其他", ".gitignore", "Git 忽略规则"),
        ],
    },
    {
        "id": "J", "name": "成员J", "type": "新增成员", "role": "数据库与项目文档",
        "duty": "数据库脚本统筹、模块文档、代码分配文档维护",
        "files": [
            ("数据库", "project/back/sql/schema.sql", "统一数据库初始化（主维护，协同 B/C/D/F）"),
            ("文档", ".doc/MODULE_FILES.md", "模块文件目录对照表"),
            ("文档", ".doc/CODE_ALLOCATION.xlsx", "10 人代码分配表（本文档）"),
            ("文档", ".doc/成员B-认证安全模块文档.docx", "成员 B 模块文档"),
            ("文档", ".doc/成员C-账户管理模块文档.docx", "成员 C 模块文档"),
            ("文档", ".doc/成员E-账单报表模块文档.docx", "成员 E 模块文档"),
        ],
    },
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ORIGINAL_FILL = PatternFill("solid", fgColor="E2EFDA")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BBBBBB")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def build_workbook():
    wb = Workbook()

    # Sheet 1: 总览
    ws1 = wb.active
    ws1.title = "成员总览"
    headers1 = ["成员", "成员类型", "角色", "职责概述", "分配文件数", "工作量占比"]
    ws1.append(headers1)
    style_header(ws1, 1, len(headers1))

    total_files = sum(len(m["files"]) for m in MEMBERS)
    original_files = sum(len(m["files"]) for m in MEMBERS if m["type"] == "原核心成员")
    new_files = total_files - original_files

    for m in MEMBERS:
        cnt = len(m["files"])
        pct = f"{cnt / total_files * 100:.1f}%"
        ws1.append([m["name"], m["type"], m["role"], m["duty"], cnt, pct])
        row = ws1.max_row
        fill = ORIGINAL_FILL if m["type"] == "原核心成员" else NEW_FILL
        for c in range(1, 7):
            ws1.cell(row=row, column=c).fill = fill
            ws1.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    ws1.append([])
    ws1.append(["合计", "", "", f"原核心 6 人共 {original_files} 个文件", original_files, f"{original_files/total_files*100:.1f}%"])
    ws1.append(["", "", "", f"新增 4 人共 {new_files} 个文件", new_files, f"{new_files/total_files*100:.1f}%"])
    ws1.append(["", "", "", f"全项目共 {total_files} 个文件", total_files, "100%"])
    auto_width(ws1)

    # Sheet 2: 文件明细
    ws2 = wb.create_sheet("文件分配明细")
    headers2 = ["成员", "成员类型", "角色", "文件分类", "文件路径", "说明"]
    ws2.append(headers2)
    style_header(ws2, 1, len(headers2))

    for m in MEMBERS:
        fill = ORIGINAL_FILL if m["type"] == "原核心成员" else NEW_FILL
        for cat, path, desc in m["files"]:
            ws2.append([m["name"], m["type"], m["role"], cat, path, desc])
            row = ws2.max_row
            for c in range(1, 7):
                ws2.cell(row=row, column=c).fill = fill
                ws2.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                ws2.cell(row=row, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    auto_width(ws2)

    # Sheet 3: 按成员分 sheet 索引
    ws3 = wb.create_sheet("按成员索引")
    ws3.append(["成员", "角色", "文件路径", "说明"])
    style_header(ws3, 1, 4)
    for m in MEMBERS:
        for _, path, desc in m["files"]:
            ws3.append([m["name"], m["role"], path, desc])
    auto_width(ws3)

    return wb, total_files, original_files, new_files


if __name__ == "__main__":
    wb, total, orig, new = build_workbook()
    out = r"d:\code\OnlineBankSystem\.doc\CODE_ALLOCATION.xlsx"
    wb.save(out)
    print(f"Generated: {out}")
    print(f"Total: {total}, Original 6: {orig} ({orig/total*100:.1f}%), New 4: {new} ({new/total*100:.1f}%)")
